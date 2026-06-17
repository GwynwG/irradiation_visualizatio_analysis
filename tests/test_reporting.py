from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from irradiation_analysis.excel_io import ImportResult, ImportSummary
from irradiation_analysis.models import (
    AbnormalEvent,
    DeviceSnapshot,
    ForecastHorizon,
    MonitoringRecord,
    MonitoringSnapshot,
    MonitoringStatus,
    QualityIssue,
    RiskResult,
    RoomRiskResult,
    SeriesForecast,
    SystemForecast,
    WarningAlert,
)
from irradiation_analysis.reporting import AnalysisReportInput, build_analysis_report


REQUIRED_SHEETS = [
    "分析摘要",
    "智能预警",
    "异常事件",
    "设备风险排名",
    "房间风险排名",
    "趋势预测",
    "清洗后的监测数据",
    "数据质量问题",
]

TABLE_SHEETS = [
    "智能预警",
    "异常事件",
    "设备风险排名",
    "房间风险排名",
    "趋势预测",
    "清洗后的监测数据",
    "数据质量问题",
]


def record(
    monitored_at: datetime,
    value: float,
    *,
    device_id: str = "R01-D01",
    monitor_type: str = "γ剂量率",
    unit: str = "μSv/h",
    warning: float = 10.0,
    control: float = 20.0,
    import_order: int = 0,
) -> MonitoringRecord:
    return MonitoringRecord(
        monitored_at=monitored_at,
        date_only=False,
        room_id=device_id.split("-")[0],
        device_id=device_id,
        monitor_type=monitor_type,
        value=value,
        unit=unit,
        warning_threshold=warning,
        control_threshold=control,
        source_file="监测数据.xlsx",
        source_sheet="监测数据",
        source_row=import_order + 2,
        import_order=import_order,
        room_name=f"{device_id.split('-')[0]}房间",
        device_name=f"{device_id}探头",
        data_source="巡检导入",
        note="fixture",
    )


def report_input() -> AnalysisReportInput:
    normal = record(datetime(2026, 6, 1, 8), 8.0, import_order=1)
    warning = record(datetime(2026, 6, 2, 8), 12.0, import_order=2)
    accident = record(
        datetime(2026, 6, 3, 8),
        22.0,
        device_id="R01-D02",
        import_order=3,
    )
    records = [normal, warning, accident]
    issue = QualityIssue(
        level="warning",
        code="unit_changed",
        message="历史单位发生变化。",
        source_file="监测数据.xlsx",
        source_sheet="监测数据",
        source_row=5,
        details={"device_id": "R01-D01", "units": ["μSv/h", "mSv/h"]},
    )
    import_result = ImportResult(
        records=records,
        issues=[issue],
        summary=ImportSummary(
            file_count=1,
            raw_rows=4,
            valid_rows=3,
            blocked_rows=1,
            exact_duplicate_rows=0,
            conflict_keys=0,
            room_count=1,
            device_count=2,
            monitor_types=("γ剂量率",),
        ),
        candidate_sheets={"监测数据.xlsx": ("监测数据",)},
    )
    snapshot = MonitoringSnapshot(
        selected_at=datetime(2026, 6, 3, 8),
        devices={
            "R01-D01": DeviceSnapshot(
                device_id="R01-D01",
                room_id="R01",
                status=MonitoringStatus.WARNING,
                latest_by_series={(warning.monitor_type, warning.unit): warning},
                most_severe_record=warning,
            ),
            "R01-D02": DeviceSnapshot(
                device_id="R01-D02",
                room_id="R01",
                status=MonitoringStatus.ACCIDENT,
                latest_by_series={(accident.monitor_type, accident.unit): accident},
                most_severe_record=accident,
            ),
        },
        room_statuses={"R01": MonitoringStatus.ACCIDENT},
    )
    event = AbnormalEvent(
        room_id="R01",
        device_id="R01-D02",
        monitor_type="γ剂量率",
        unit="μSv/h",
        started_at=datetime(2026, 6, 3, 8),
        ended_at=datetime(2026, 6, 4, 8),
        highest_status=MonitoringStatus.ACCIDENT,
        peak_value=22.0,
        peak_time=datetime(2026, 6, 3, 8),
        record_count=1,
        duration_days=1.0,
        source_records=(accident,),
        reasons=("最高状态为事故级",),
    )
    device_risk = RiskResult(
        room_id="R01",
        device_id="R01-D01",
        score=66.5,
        status=MonitoringStatus.WARNING,
        component_scores={"severity": 60.0, "exceedance": 47.0},
        events=(event,),
        growth_signals=(),
        record_count=2,
        reasons=("当前状态为预警", "abnormal events=1"),
    )
    room_risk = RoomRiskResult(
        room_id="R01",
        score=78.0,
        max_device_score=66.5,
        abnormal_device_ratio=0.5,
        duration_score=20.0,
        device_count=2,
        abnormal_device_count=1,
        event_count=1,
        longest_event_duration_days=1.0,
        reasons=("max_device_score=66.5",),
    )
    series_forecast = SeriesForecast(
        room_id="R01",
        device_id="R01-D01",
        monitor_type="γ剂量率",
        unit="μSv/h",
        horizon=ForecastHorizon.DAYS_7,
        predicted_at=datetime(2026, 6, 10, 8),
        predicted_value=13.5,
        predicted_status=MonitoringStatus.WARNING,
        warning_threshold=10.0,
        control_threshold=20.0,
        method="线性趋势",
        sample_count=4,
        training_start=datetime(2026, 6, 1, 8),
        training_end=datetime(2026, 6, 3, 8),
        confidence="中",
        explanation="样本有限，仅供趋势参考。",
    )
    system_forecast = SystemForecast(
        horizon=ForecastHorizon.DAYS_7,
        series_forecasts=(series_forecast,),
        device_statuses={
            "R01-D01": MonitoringStatus.WARNING,
            "R01-D02": MonitoringStatus.ACCIDENT,
        },
        normal_devices=0,
        warning_devices=1,
        accident_devices=1,
        no_data_devices=198,
        summary="forecasted_series=1",
        reasons=("预测不替代现场复核",),
    )
    warning_alert = WarningAlert(
        room_id="R01",
        device_id="R01-D01",
        monitor_type="γ剂量率",
        unit="μSv/h",
        rule_code="current_warning",
        level="预警",
        score=82.0,
        triggered_at=datetime(2026, 6, 2, 8),
        current_value=12.0,
        warning_threshold=10.0,
        control_threshold=20.0,
        evidence="当前值 12 已达到或超过预警值 10。",
        recommended_action="安排复测并确认近期作业、屏蔽和仪器校准状态。",
    )
    return AnalysisReportInput(
        import_result=import_result,
        snapshot=snapshot,
        events=(event,),
        device_risks=(device_risk,),
        room_risks=(room_risk,),
        series_forecasts=(series_forecast,),
        system_forecast=system_forecast,
        generated_at=datetime(2026, 6, 16, 9, 30),
        warning_alerts=(warning_alert,),
    )


def workbook_from_report():
    content = build_analysis_report(report_input())
    return load_workbook(BytesIO(content), data_only=True)


def column_for_header(worksheet, header: str, header_row: int = 1) -> int:
    headers = [cell.value for cell in worksheet[header_row]]
    return headers.index(header) + 1


def assert_fill_suffix(cell, suffix: str) -> None:
    assert cell.fill.fill_type == "solid"
    assert cell.fill.fgColor.rgb is not None
    assert cell.fill.fgColor.rgb.upper().endswith(suffix)


def assert_column_format(
    worksheet,
    header: str,
    number_format: str,
    *,
    header_row: int = 1,
) -> None:
    column = column_for_header(worksheet, header, header_row)
    assert worksheet.cell(row=header_row + 1, column=column).number_format == number_format


def summary_pairs(worksheet) -> dict[str, object]:
    pairs = {}
    for row in worksheet.iter_rows(values_only=True):
        for label, value in ((row[0], row[1]), (row[2], row[3])):
            if label:
                pairs[label] = value
    return pairs


def test_analysis_report_contains_all_required_sheets():
    workbook = workbook_from_report()

    assert workbook.sheetnames == REQUIRED_SHEETS

    workbook.close()


def test_report_formats_tables_statuses_and_data_columns():
    workbook = workbook_from_report()

    events_sheet = workbook["异常事件"]
    alerts_sheet = workbook["智能预警"]
    device_risk_sheet = workbook["设备风险排名"]
    room_risk_sheet = workbook["房间风险排名"]
    forecast_sheet = workbook["趋势预测"]
    cleaned_sheet = workbook["清洗后的监测数据"]

    for sheet_name in TABLE_SHEETS:
        worksheet = workbook[sheet_name]
        assert worksheet.auto_filter.ref is not None
        expected_freeze = "A9" if sheet_name == "数据质量问题" else "A2"
        assert worksheet.freeze_panes == expected_freeze

    event_status_column = column_for_header(events_sheet, "最高状态")
    risk_status_column = column_for_header(device_risk_sheet, "当前状态")
    forecast_status_column = column_for_header(forecast_sheet, "预测状态")
    assert_fill_suffix(events_sheet.cell(row=2, column=event_status_column), "DC2626")
    assert_fill_suffix(device_risk_sheet.cell(row=2, column=risk_status_column), "F59E0B")
    assert_fill_suffix(forecast_sheet.cell(row=2, column=forecast_status_column), "F59E0B")

    assert_column_format(alerts_sheet, "触发时间", "yyyy-mm-dd hh:mm")
    for header in ("开始时间", "结束时间", "峰值时间"):
        assert_column_format(events_sheet, header, "yyyy-mm-dd hh:mm")
    for header in ("预测时间", "训练开始", "训练结束"):
        assert_column_format(forecast_sheet, header, "yyyy-mm-dd hh:mm")
    assert_column_format(cleaned_sheet, "监测时间", "yyyy-mm-dd hh:mm")

    for header in ("评分", "当前或预测值", "预警值", "控制标准"):
        assert_column_format(alerts_sheet, header, "0.000")
    for header in ("峰值", "持续天数"):
        assert_column_format(events_sheet, header, "0.000")
    for header in ("风险评分", "严重度分", "超限分", "持续分", "趋势分", "复发分"):
        assert_column_format(device_risk_sheet, header, "0.000")
    for header in ("风险评分", "最高设备评分", "持续分", "最长事件持续天数"):
        assert_column_format(room_risk_sheet, header, "0.000")
    assert_column_format(room_risk_sheet, "异常设备比例", "0.00%")
    for header in ("预测值", "预警值", "控制标准"):
        assert_column_format(forecast_sheet, header, "0.000")
    for header in ("监测值", "预警值", "控制标准"):
        assert_column_format(cleaned_sheet, header, "0.000")

    workbook.close()


def test_summary_sheet_includes_required_high_level_content():
    workbook = workbook_from_report()

    pairs = summary_pairs(workbook["分析摘要"])
    values = [
        value
        for row in workbook["分析摘要"].iter_rows(values_only=True)
        for value in row
        if value is not None
    ]

    assert pairs["导入文件数"] == 1
    assert pairs["原始行数"] == 4
    assert pairs["有效行数"] == 3
    assert pairs["阻断行数"] == 1
    assert pairs["完全重复行数"] == 0
    assert pairs["冲突键数"] == 0
    assert pairs["房间数"] == 1
    assert pairs["设备数"] == 2
    assert pairs["监测类型"] == "γ剂量率"
    assert pairs["质量分"] == 82.2
    assert pairs["质量等级"] == "可用"

    assert pairs["设备正常数"] == 0
    assert pairs["设备预警数"] == 1
    assert pairs["设备事故级数"] == 1
    assert pairs["设备无数据数"] == 0
    assert pairs["房间正常数"] == 0
    assert pairs["房间预警数"] == 0
    assert pairs["房间事故级数"] == 1
    assert pairs["房间无数据数"] == 0

    assert pairs["异常事件数"] == 1
    assert pairs["设备风险条目数"] == 1
    assert pairs["房间风险条目数"] == 1
    assert pairs["最高设备风险"] == "R01-D01（66.5）"
    assert pairs["最高房间风险"] == "R01（78.0）"
    assert pairs["智能预警数"] == 1
    assert pairs["事故级预警数"] == 0
    assert pairs["预测触发预警数"] == 0
    assert pairs["预测序列数"] == 1
    assert pairs["预测正常设备数"] == 0
    assert pairs["预测预警设备数"] == 1
    assert pairs["预测事故级设备数"] == 1
    assert pairs["预测无数据设备数"] == 198
    assert pairs["系统预测摘要"] == "forecasted_series=1"
    assert any("预测结果仅供趋势研判参考" in str(value) for value in values)

    workbook.close()


def test_report_includes_source_fields_forecast_context_and_quality_details():
    workbook = workbook_from_report()

    cleaned_headers = [cell.value for cell in workbook["清洗后的监测数据"][1]]
    alert_sheet = workbook["智能预警"]
    alert_headers = [cell.value for cell in alert_sheet[1]]
    forecast_sheet = workbook["趋势预测"]
    forecast_headers = [cell.value for cell in forecast_sheet[1]]
    quality_sheet = workbook["数据质量问题"]
    quality_headers = [cell.value for cell in quality_sheet[8]]
    forecast_values = [
        value
        for row in forecast_sheet.iter_rows(values_only=True)
        for value in row
        if value is not None
    ]
    alert_values = [
        value
        for row in alert_sheet.iter_rows(values_only=True)
        for value in row
        if value is not None
    ]
    quality_values = [
        value
        for row in quality_sheet.iter_rows(values_only=True)
        for value in row
        if value is not None
    ]

    assert {"来源文件", "来源工作表", "来源行号"}.issubset(cleaned_headers)
    assert {"触发规则", "证据", "建议动作"}.issubset(alert_headers)
    assert any("current_warning" in str(value) for value in alert_values)
    assert any("安排复测" in str(value) for value in alert_values)
    assert {"预测方法", "样本数", "置信度", "说明"}.issubset(forecast_headers)
    assert any("预测结果仅供趋势研判参考" in str(value) for value in forecast_values)
    assert {"来源文件", "来源工作表", "来源行号", "详情"}.issubset(quality_headers)
    assert quality_sheet["A1"].value == "数据质量概览"
    assert quality_sheet["B2"].value == 82.2
    assert quality_sheet["D2"].value == "可用"
    assert any("unit_changed" in str(value) for value in quality_values)
    assert any("R01-D01" in str(value) for value in quality_values)

    workbook.close()
