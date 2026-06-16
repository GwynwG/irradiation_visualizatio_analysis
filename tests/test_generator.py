from __future__ import annotations

from collections import Counter
from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from irradiation_analysis.analytics import find_growth_signals
from irradiation_analysis.excel_io import (
    OPTIONAL_COLUMNS,
    REQUIRED_COLUMNS,
    UploadedWorkbook,
    import_workbooks,
)
from irradiation_analysis.generator import (
    DEFAULT_MONITOR_TYPES,
    MonitorTypeConfig,
    SimulationConfig,
    build_blank_template,
    build_prefilled_template,
    generate_simulated_workbooks,
)
from irradiation_analysis.snapshots import all_device_ids
from irradiation_analysis.models import MonitoringStatus
from irradiation_analysis.status import classify_record


DATA_SHEET = "监测数据"


def workbook_from_bytes(content: bytes):
    return load_workbook(BytesIO(content), data_only=True)


def test_blank_template_has_required_headers_and_frozen_pane():
    workbook = workbook_from_bytes(build_blank_template())
    worksheet = workbook[DATA_SHEET]

    headers = [cell.value for cell in worksheet[1]]

    assert headers == list(REQUIRED_COLUMNS + OPTIONAL_COLUMNS)
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:L2"
    assert worksheet.column_dimensions["A"].width >= 18
    assert worksheet.column_dimensions["E"].width >= 12
    assert worksheet["A2"].number_format == "yyyy-mm-dd hh:mm"
    assert worksheet["E2"].number_format == "0.000"
    assert "设备阈值配置" in workbook.sheetnames

    workbook.close()


def test_prefilled_template_contains_two_hundred_devices_per_monitor_type():
    monitor_types = (
        MonitorTypeConfig("γ剂量率", "μSv/h", 10.0, 20.0),
        MonitorTypeConfig("中子剂量率", "μSv/h", 5.0, 12.0),
    )
    config = SimulationConfig(
        start=datetime(2026, 6, 1, 8),
        end=datetime(2026, 6, 1, 8),
        monitor_types=monitor_types,
        seed=123,
    )

    content = build_prefilled_template(config)
    workbook = workbook_from_bytes(content)
    worksheet = workbook[DATA_SHEET]
    rows = list(
        worksheet.iter_rows(
            min_row=2,
            values_only=True,
        )
    )
    monitor_type_counts = Counter(row[3] for row in rows if row[0] is not None)
    device_ids = {row[2] for row in rows if row[0] is not None}

    assert len(rows) == 400
    assert monitor_type_counts == {"γ剂量率": 200, "中子剂量率": 200}
    assert device_ids == set(all_device_ids())

    result = import_workbooks([UploadedWorkbook("prefilled.xlsx", content)])
    assert result.summary.valid_rows == 400
    assert result.summary.blocked_rows == 0

    workbook.close()


def test_simulation_is_reproducible_and_importable():
    config = SimulationConfig(
        start=datetime(2026, 6, 1),
        end=datetime(2026, 6, 6),
        sampling_hours=24,
        warning_ratio=0.05,
        accident_ratio=0.03,
        rapid_growth_ratio=0.04,
        event_duration=3,
        seed=20260616,
    )

    first = generate_simulated_workbooks(config)
    second = generate_simulated_workbooks(config)

    assert [(workbook.filename, workbook.content) for workbook in first] == [
        (workbook.filename, workbook.content) for workbook in second
    ]

    result = import_workbooks(first)

    assert len(first) == 1
    assert first[0].filename == "simulated_monitoring_20260601_20260606.xlsx"
    assert result.summary.blocked_rows == 0
    assert result.summary.valid_rows > 0
    assert set(DEFAULT_MONITOR_TYPES).issubset(result.summary.monitor_types)


def test_simulation_includes_requested_statuses_and_consistent_thresholds():
    config = SimulationConfig(
        start=datetime(2026, 6, 1),
        end=datetime(2026, 6, 6),
        sampling_hours=24,
        warning_ratio=0.05,
        accident_ratio=0.03,
        rapid_growth_ratio=0.04,
        event_duration=3,
        seed=42,
    )
    result = import_workbooks(generate_simulated_workbooks(config))
    statuses = {classify_record(record) for record in result.records}

    assert statuses == {
        MonitoringStatus.NORMAL,
        MonitoringStatus.WARNING,
        MonitoringStatus.ACCIDENT,
    }
    assert all(
        0 < record.warning_threshold < record.control_threshold
        for record in result.records
    )
    assert find_growth_signals(result.records)


def test_unknown_output_mode_raises_clear_error():
    config = SimulationConfig(
        start=datetime(2026, 6, 1),
        end=datetime(2026, 6, 1),
        output_mode="weekly",
    )

    with pytest.raises(ValueError, match="output_mode"):
        generate_simulated_workbooks(config)
