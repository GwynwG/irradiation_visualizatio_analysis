from datetime import date, datetime
from io import BytesIO
import zipfile

import pytest
from openpyxl import Workbook

from irradiation_analysis.excel_io import (
    OPTIONAL_COLUMNS,
    REQUIRED_COLUMNS,
    UploadedWorkbook,
    import_workbooks,
)


def workbook_bytes(
    sheets: list[tuple[str, list[str], list[list[object]], list[list[object]] | None]],
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for name, headers, rows, preamble in sheets:
        worksheet = workbook.create_sheet(name)
        for preamble_row in preamble or []:
            worksheet.append(preamble_row)
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def uploaded(
    filename: str,
    rows: list[list[object]],
    *,
    sheet: str = "监测数据",
    headers: list[str] | None = None,
    preamble: list[list[object]] | None = None,
) -> UploadedWorkbook:
    return UploadedWorkbook(
        filename,
        workbook_bytes(
            [
                (
                    sheet,
                    headers or list(REQUIRED_COLUMNS + OPTIONAL_COLUMNS),
                    rows,
                    preamble,
                )
            ]
        ),
    )


def row(
    monitored_at: object = "2026-06-15",
    room_id: object = "R01",
    device_id: object = "R01-D01",
    monitor_type: object = "γ剂量率",
    value: object = 1.0,
    unit: object = "μSv/h",
    warning: object = 10.0,
    control: object = 20.0,
    room_name: object = "",
    device_name: object = "",
    data_source: object = "",
    note: object = "",
) -> list[object]:
    return [
        monitored_at,
        room_id,
        device_id,
        monitor_type,
        value,
        unit,
        warning,
        control,
        room_name,
        device_name,
        data_source,
        note,
    ]


def issue_codes(result) -> list[str]:
    return [issue.code for issue in result.issues]


def test_imports_long_table_and_normalizes_dates_and_text_fields():
    headers = [
        "备注",
        "设备ID",
        "监测时间",
        "房间ID",
        "监测值",
        "单位",
        "监测类型",
        "控制标准",
        "预警值",
        "设备名称",
        "额外列",
        "房间名称",
        "数据来源",
    ]
    rows = [
        [
            "  每日记录  ",
            " R01-D01 ",
            date(2026, 6, 1),
            " R01 ",
            1.25,
            " μSv/h ",
            " γ剂量率 ",
            20,
            10,
            " 设备甲 ",
            "ignored",
            " 一号房 ",
            " 系统A ",
        ],
        [
            "",
            "R01-D02",
            datetime(2026, 6, 2, 14, 30, 15),
            "R01",
            2.5,
            "μSv/h",
            "γ剂量率",
            20,
            10,
            "",
            "ignored",
            "",
            "",
        ],
    ]

    result = import_workbooks(
        [
            uploaded(
                "long.xlsx",
                rows,
                headers=headers,
                preamble=[["长期监测导出"], []],
            )
        ]
    )

    assert result.candidate_sheets == {"long.xlsx": ("监测数据",)}
    assert len(result.records) == 2
    first, second = result.records
    assert first.monitored_at == datetime(2026, 6, 1)
    assert first.date_only is True
    assert first.room_id == "R01"
    assert first.device_id == "R01-D01"
    assert first.monitor_type == "γ剂量率"
    assert first.unit == "μSv/h"
    assert first.room_name == "一号房"
    assert first.device_name == "设备甲"
    assert first.data_source == "系统A"
    assert first.note == "每日记录"
    assert first.source_row == 4
    assert second.monitored_at == datetime(2026, 6, 2, 14, 30, 15)
    assert second.date_only is False


def test_finds_required_headers_after_first_twenty_rows():
    result = import_workbooks(
        [
            uploaded(
                "late-header.xlsx",
                [row()],
                preamble=[[f"note {index}"] for index in range(25)],
            )
        ]
    )

    assert len(result.records) == 1
    assert result.records[0].source_row == 27
    assert result.summary.valid_rows == 1
    assert "required_columns_missing" not in issue_codes(result)


def test_parses_supported_string_date_and_datetime_formats():
    result = import_workbooks(
        [
            uploaded(
                "dates.xlsx",
                [
                    row(monitored_at="2026-06-01", device_id="R01-D01"),
                    row(monitored_at="2026/06/02", device_id="R01-D02"),
                    row(
                        monitored_at="2026-06-03 04:05",
                        device_id="R01-D03",
                    ),
                    row(
                        monitored_at="2026-06-04 04:05:06",
                        device_id="R01-D04",
                    ),
                    row(
                        monitored_at="2026-06-05T04:05:06",
                        device_id="R01-D05",
                    ),
                ],
            )
        ]
    )

    assert [record.date_only for record in result.records] == [
        True,
        True,
        False,
        False,
        False,
    ]
    assert result.records[-1].monitored_at == datetime(2026, 6, 5, 4, 5, 6)


def test_deduplicates_exact_rows_then_selects_latest_conflicting_version():
    result = import_workbooks(
        [
            uploaded(
                "conflicts.xlsx",
                [
                    row(value=10),
                    row(value=10),
                    row(value=12),
                ],
            )
        ]
    )

    assert len(result.records) == 1
    assert result.records[0].value == 12
    assert result.records[0].import_order == 2
    assert result.summary.raw_rows == 3
    assert result.summary.valid_rows == 3
    assert result.summary.exact_duplicate_rows == 1
    assert result.summary.conflict_keys == 1

    issue = next(issue for issue in result.issues if issue.code == "conflicting_record")
    assert issue.details["selected_source"]["row"] == 4
    assert [version["value"] for version in issue.details["versions"]] == [10.0, 12.0]
    assert all(
        {"source_file", "source_sheet", "source_row", "value", "unit", "warning_threshold", "control_threshold"}
        <= version.keys()
        for version in issue.details["versions"]
    )


def test_invalid_threshold_order_blocks_row():
    result = import_workbooks(
        [uploaded("thresholds.xlsx", [row(warning=20, control=10)])]
    )

    assert result.records == []
    assert result.summary.raw_rows == 1
    assert result.summary.valid_rows == 0
    assert result.summary.blocked_rows == 1
    assert issue_codes(result) == ["invalid_threshold_order"]


def test_invalid_rows_are_blocked_with_specific_issue_codes():
    result = import_workbooks(
        [
            uploaded(
                "invalid.xlsx",
                [
                    row(room_id="R00", device_id="R00-D01"),
                    row(room_id="R01", device_id="R02-D01"),
                    row(unit="   "),
                    row(monitored_at="not-a-date"),
                    row(value=True),
                    row(warning="nan"),
                    row(control="inf"),
                ],
            )
        ]
    )

    assert result.records == []
    assert result.summary.raw_rows == 7
    assert result.summary.valid_rows == 0
    assert result.summary.blocked_rows == 7
    assert issue_codes(result) == [
        "invalid_identifier",
        "invalid_identifier",
        "required_value_missing",
        "invalid_monitored_at",
        "invalid_value",
        "invalid_warning_threshold",
        "invalid_control_threshold",
    ]
    identifier_issues = [
        issue for issue in result.issues if issue.code == "invalid_identifier"
    ]
    assert all(issue.details["errors"] for issue in identifier_issues)
    assert "prefix" in identifier_issues[1].details["errors"][0]


def test_multiple_candidate_sheets_require_explicit_selection():
    content = workbook_bytes(
        [
            ("日监测", list(REQUIRED_COLUMNS), [row()[:8]], None),
            (
                "月监测",
                list(REQUIRED_COLUMNS),
                [row(device_id="R01-D02", value=2)[:8]],
                None,
            ),
        ]
    )
    workbook = UploadedWorkbook("multi.xlsx", content)

    unselected = import_workbooks([workbook])

    assert unselected.records == []
    assert unselected.candidate_sheets == {"multi.xlsx": ("日监测", "月监测")}
    assert unselected.summary.raw_rows == 0
    assert unselected.summary.blocked_rows == 0
    assert issue_codes(unselected) == ["sheet_selection_required"]

    selected = import_workbooks(
        [workbook],
        selected_sheets={"multi.xlsx": "月监测"},
    )

    assert len(selected.records) == 1
    assert selected.records[0].source_sheet == "月监测"
    assert selected.records[0].device_id == "R01-D02"


def test_workbook_without_matching_sheet_reports_missing_columns():
    workbook = UploadedWorkbook(
        "wrong.xlsx",
        workbook_bytes([("说明", ["房间ID", "设备ID"], [["R01", "R01-D01"]], None)]),
    )

    result = import_workbooks([workbook])

    assert result.records == []
    assert result.candidate_sheets == {"wrong.xlsx": ()}
    assert result.summary.raw_rows == 0
    assert issue_codes(result) == ["required_columns_missing"]


def test_upload_and_row_order_select_the_latest_conflict():
    first = uploaded("first.xlsx", [row(value=10), row(value=11)])
    second = uploaded("second.xlsx", [row(value=12)])

    forward = import_workbooks([first, second])
    reverse = import_workbooks([second, first])

    assert forward.records[0].value == 12
    assert forward.records[0].source_file == "second.xlsx"
    assert forward.records[0].import_order == 2
    assert reverse.records[0].value == 11
    assert reverse.records[0].source_file == "first.xlsx"
    assert reverse.records[0].import_order == 2


def test_audits_historical_unit_and_threshold_changes_without_blocking():
    result = import_workbooks(
        [
            uploaded(
                "history.xlsx",
                [
                    row(monitored_at="2026-06-01", unit="μSv/h", warning=10, control=20),
                    row(monitored_at="2026-06-02", unit="nSv/h", warning=10, control=20),
                    row(monitored_at="2026-06-03", unit="nSv/h", warning=12, control=24),
                ],
            )
        ]
    )

    assert len(result.records) == 3
    audit_issues = [
        issue
        for issue in result.issues
        if issue.code in {"unit_changed", "threshold_changed"}
    ]
    assert [issue.code for issue in audit_issues] == [
        "unit_changed",
        "threshold_changed",
    ]
    assert all(issue.level == "warning" for issue in audit_issues)
    assert result.summary.blocked_rows == 0


def test_summary_counts_raw_valid_blocked_entities_and_sorted_types():
    first = uploaded(
        "one.xlsx",
        [
            row(
                room_id="R02",
                device_id="R02-D02",
                monitor_type="表面污染",
            ),
            row(
                room_id="R01",
                device_id="R01-D01",
                monitor_type="γ剂量率",
            ),
            row(
                room_id="R01",
                device_id="R01-D02",
                monitor_type="γ剂量率",
                value="nan",
            ),
            [None] * len(REQUIRED_COLUMNS + OPTIONAL_COLUMNS),
        ],
    )
    second = uploaded(
        "two.xlsx",
        [
            row(
                monitored_at="2026-06-16",
                room_id="R01",
                device_id="R01-D01",
                monitor_type="γ剂量率",
            )
        ],
    )

    result = import_workbooks([first, second])

    assert result.summary.file_count == 2
    assert result.summary.raw_rows == 4
    assert result.summary.valid_rows == 3
    assert result.summary.blocked_rows == 1
    assert result.summary.room_count == 2
    assert result.summary.device_count == 2
    assert result.summary.monitor_types == ("γ剂量率", "表面污染")


def test_corrupt_workbook_reports_read_error_instead_of_raising():
    result = import_workbooks([UploadedWorkbook("broken.xlsx", b"not an xlsx")])

    assert result.records == []
    assert result.summary.file_count == 1
    assert issue_codes(result) == ["workbook_read_error"]


def test_lazily_parsed_corrupt_sheet_reports_read_error():
    valid_content = workbook_bytes(
        [("监测数据", list(REQUIRED_COLUMNS), [row()[:8]], None)]
    )
    source = BytesIO(valid_content)
    corrupted = BytesIO()
    with zipfile.ZipFile(source) as input_zip:
        with zipfile.ZipFile(corrupted, "w") as output_zip:
            for entry in input_zip.infolist():
                content = input_zip.read(entry.filename)
                if entry.filename == "xl/worksheets/sheet1.xml":
                    dimension_end = content.index(b"/>", content.index(b"<dimension")) + 2
                    content = content[:dimension_end] + b"<sheetData><row"
                output_zip.writestr(entry, content)

    result = import_workbooks(
        [UploadedWorkbook("lazy-broken.xlsx", corrupted.getvalue())]
    )

    assert result.records == []
    assert issue_codes(result) == ["workbook_read_error"]


@pytest.mark.parametrize("selected_sheet", ["missing", ""])
def test_invalid_selected_sheet_still_requires_a_valid_selection(selected_sheet):
    workbook = UploadedWorkbook(
        "multi.xlsx",
        workbook_bytes(
            [
                ("A", list(REQUIRED_COLUMNS), [row()[:8]], None),
                ("B", list(REQUIRED_COLUMNS), [row()[:8]], None),
            ]
        ),
    )

    result = import_workbooks(
        [workbook],
        selected_sheets={"multi.xlsx": selected_sheet},
    )

    assert result.records == []
    assert issue_codes(result) == ["sheet_selection_required"]
