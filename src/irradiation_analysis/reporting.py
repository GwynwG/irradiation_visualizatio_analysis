from __future__ import annotations

from collections import Counter
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from irradiation_analysis.excel_io import ImportResult
from irradiation_analysis.models import (
    AbnormalEvent,
    MonitoringSnapshot,
    MonitoringRecord,
    MonitoringStatus,
    QualityIssue,
    RiskResult,
    RoomRiskResult,
    SeriesForecast,
    SystemForecast,
)


SHEET_NAMES = (
    "分析摘要",
    "异常事件",
    "设备风险排名",
    "房间风险排名",
    "趋势预测",
    "清洗后的监测数据",
    "数据质量问题",
)

FORECAST_DISCLAIMER = (
    "预测结果仅供趋势研判参考，不替代现场复核、仪器校准或安全处置决策。"
)

DATETIME_FORMAT = "yyyy-mm-dd hh:mm"
NUMBER_FORMAT = "0.000"
INTEGER_FORMAT = "0"
PERCENT_FORMAT = "0.00%"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="E2E8F0")
STATUS_FILLS = {
    MonitoringStatus.NO_DATA: PatternFill(fill_type="solid", fgColor="94A3B8"),
    MonitoringStatus.NORMAL: PatternFill(fill_type="solid", fgColor="22C55E"),
    MonitoringStatus.WARNING: PatternFill(fill_type="solid", fgColor="F59E0B"),
    MonitoringStatus.ACCIDENT: PatternFill(fill_type="solid", fgColor="DC2626"),
}


@dataclass(frozen=True)
class AnalysisReportInput:
    import_result: ImportResult
    snapshot: MonitoringSnapshot
    events: Sequence[AbnormalEvent]
    device_risks: Sequence[RiskResult]
    room_risks: Sequence[RoomRiskResult]
    series_forecasts: Sequence[SeriesForecast]
    system_forecast: SystemForecast | None
    generated_at: datetime | None = None


def build_analysis_report(report_input: AnalysisReportInput) -> bytes:
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = SHEET_NAMES[0]
    worksheets = {SHEET_NAMES[0]: first_sheet}
    for sheet_name in SHEET_NAMES[1:]:
        worksheets[sheet_name] = workbook.create_sheet(sheet_name)

    _write_summary_sheet(worksheets["分析摘要"], report_input)
    _write_events_sheet(worksheets["异常事件"], report_input.events)
    _write_device_risks_sheet(worksheets["设备风险排名"], report_input.device_risks)
    _write_room_risks_sheet(worksheets["房间风险排名"], report_input.room_risks)
    _write_forecasts_sheet(worksheets["趋势预测"], report_input.series_forecasts)
    _write_cleaned_data_sheet(
        worksheets["清洗后的监测数据"],
        report_input.import_result.records,
    )
    _write_quality_issues_sheet(
        worksheets["数据质量问题"],
        report_input.import_result.issues,
    )

    generated_at = report_input.generated_at or datetime.now()
    workbook.properties.creator = "irradiation-analysis"
    workbook.properties.lastModifiedBy = "irradiation-analysis"
    workbook.properties.created = generated_at
    workbook.properties.modified = generated_at

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _write_summary_sheet(
    worksheet: Worksheet,
    report_input: AnalysisReportInput,
) -> None:
    worksheet["A1"] = "辐照监测分析报告"
    worksheet["A1"].font = Font(bold=True, size=16)
    worksheet.merge_cells("A1:D1")
    worksheet.freeze_panes = "A3"

    generated_at = report_input.generated_at or datetime.now()
    summary = report_input.import_result.summary
    device_status_counts = _status_counts(
        device.status for device in report_input.snapshot.devices.values()
    )
    room_status_counts = _status_counts(report_input.snapshot.room_statuses.values())
    top_device_risk = _top_risk_label(report_input.device_risks)
    top_room_risk = _top_room_risk_label(report_input.room_risks)
    system_forecast = report_input.system_forecast

    rows: list[tuple[str, object, str, object]] = [
        ("生成时间", generated_at, "快照时间", report_input.snapshot.selected_at),
        ("", "", "", ""),
        ("导入摘要", "", "", ""),
        ("导入文件数", summary.file_count, "原始行数", summary.raw_rows),
        ("有效行数", summary.valid_rows, "阻断行数", summary.blocked_rows),
        ("完全重复行数", summary.exact_duplicate_rows, "冲突键数", summary.conflict_keys),
        ("房间数", summary.room_count, "设备数", summary.device_count),
        ("监测类型", "、".join(summary.monitor_types), "", ""),
        ("", "", "", ""),
        ("快照状态统计", "", "", ""),
        (
            "设备正常数",
            device_status_counts[MonitoringStatus.NORMAL],
            "房间正常数",
            room_status_counts[MonitoringStatus.NORMAL],
        ),
        (
            "设备预警数",
            device_status_counts[MonitoringStatus.WARNING],
            "房间预警数",
            room_status_counts[MonitoringStatus.WARNING],
        ),
        (
            "设备事故级数",
            device_status_counts[MonitoringStatus.ACCIDENT],
            "房间事故级数",
            room_status_counts[MonitoringStatus.ACCIDENT],
        ),
        (
            "设备无数据数",
            device_status_counts[MonitoringStatus.NO_DATA],
            "房间无数据数",
            room_status_counts[MonitoringStatus.NO_DATA],
        ),
        ("", "", "", ""),
        ("风险和异常", "", "", ""),
        ("异常事件数", len(report_input.events), "设备风险条目数", len(report_input.device_risks)),
        ("房间风险条目数", len(report_input.room_risks), "最高设备风险", top_device_risk),
        ("最高房间风险", top_room_risk, "", ""),
        ("", "", "", ""),
        ("趋势预测", "", "", ""),
        (
            "预测序列数",
            len(report_input.series_forecasts),
            "预测周期",
            _enum_value(system_forecast.horizon) if system_forecast else "",
        ),
        (
            "预测正常设备数",
            system_forecast.normal_devices if system_forecast else "",
            "预测预警设备数",
            system_forecast.warning_devices if system_forecast else "",
        ),
        (
            "预测事故级设备数",
            system_forecast.accident_devices if system_forecast else "",
            "预测无数据设备数",
            system_forecast.no_data_devices if system_forecast else "",
        ),
        ("系统预测摘要", system_forecast.summary if system_forecast else "", "", ""),
        ("预测说明", FORECAST_DISCLAIMER, "", ""),
    ]

    for row in rows:
        worksheet.append(row)

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_index in (4, 11, 17, 22):
        for cell in worksheet[row_index]:
            cell.font = Font(bold=True)
            cell.fill = SECTION_FILL
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
    for row_index in (2,):
        worksheet.cell(row=row_index, column=2).number_format = DATETIME_FORMAT
        worksheet.cell(row=row_index, column=4).number_format = DATETIME_FORMAT
    _set_widths(worksheet, (18, 32, 18, 32))


def _write_events_sheet(
    worksheet: Worksheet,
    events: Sequence[AbnormalEvent],
) -> None:
    headers = (
        "房间ID",
        "设备ID",
        "监测类型",
        "单位",
        "开始时间",
        "结束时间",
        "最高状态",
        "峰值",
        "峰值时间",
        "记录数",
        "持续天数",
        "原因",
        "来源文件",
        "来源工作表",
        "来源行号",
    )
    _write_headers(worksheet, headers)
    for event in events:
        source_file, source_sheet, source_row = _source_columns(event.source_records)
        worksheet.append(
            (
                event.room_id,
                event.device_id,
                event.monitor_type,
                event.unit,
                event.started_at,
                event.ended_at,
                _status_label(event.highest_status),
                event.peak_value,
                event.peak_time,
                event.record_count,
                event.duration_days,
                _join_reasons(event.reasons),
                source_file,
                source_sheet,
                source_row,
            )
        )
        _style_status_cell(
            worksheet.cell(row=worksheet.max_row, column=7),
            event.highest_status,
        )
    _finalize_table(
        worksheet,
        len(headers),
        datetime_columns=(5, 6, 9),
        numeric_columns=(8, 11),
        integer_columns=(10,),
    )


def _write_device_risks_sheet(
    worksheet: Worksheet,
    risks: Sequence[RiskResult],
) -> None:
    headers = (
        "排名",
        "房间ID",
        "设备ID",
        "风险评分",
        "当前状态",
        "记录数",
        "异常事件数",
        "增长信号数",
        "严重度分",
        "超限分",
        "持续分",
        "趋势分",
        "复发分",
        "原因",
    )
    _write_headers(worksheet, headers)
    for rank, risk in enumerate(risks, start=1):
        worksheet.append(
            (
                rank,
                risk.room_id,
                risk.device_id,
                risk.score,
                _status_label(risk.status),
                risk.record_count,
                len(risk.events),
                len(risk.growth_signals),
                risk.component_scores.get("severity", 0.0),
                risk.component_scores.get("exceedance", 0.0),
                risk.component_scores.get("duration", 0.0),
                risk.component_scores.get("trend", 0.0),
                risk.component_scores.get("recurrence", 0.0),
                _join_reasons(risk.reasons),
            )
        )
        _style_status_cell(
            worksheet.cell(row=worksheet.max_row, column=5),
            risk.status,
        )
    _finalize_table(
        worksheet,
        len(headers),
        numeric_columns=(4, 9, 10, 11, 12, 13),
        integer_columns=(1, 6, 7, 8),
    )


def _write_room_risks_sheet(
    worksheet: Worksheet,
    risks: Sequence[RoomRiskResult],
) -> None:
    headers = (
        "排名",
        "房间ID",
        "风险评分",
        "最高设备评分",
        "异常设备比例",
        "持续分",
        "设备数",
        "异常设备数",
        "事件数",
        "最长事件持续天数",
        "原因",
    )
    _write_headers(worksheet, headers)
    for rank, risk in enumerate(risks, start=1):
        worksheet.append(
            (
                rank,
                risk.room_id,
                risk.score,
                risk.max_device_score,
                risk.abnormal_device_ratio,
                risk.duration_score,
                risk.device_count,
                risk.abnormal_device_count,
                risk.event_count,
                risk.longest_event_duration_days,
                _join_reasons(risk.reasons),
            )
        )
    _finalize_table(
        worksheet,
        len(headers),
        numeric_columns=(3, 4, 6, 10),
        integer_columns=(1, 7, 8, 9),
        percent_columns=(5,),
    )


def _write_forecasts_sheet(
    worksheet: Worksheet,
    forecasts: Sequence[SeriesForecast],
) -> None:
    headers = (
        "房间ID",
        "设备ID",
        "监测类型",
        "单位",
        "预测周期",
        "预测时间",
        "预测值",
        "预测状态",
        "预警值",
        "控制标准",
        "预测方法",
        "样本数",
        "训练开始",
        "训练结束",
        "置信度",
        "说明",
        "免责声明",
    )
    _write_headers(worksheet, headers)
    for forecast in forecasts:
        worksheet.append(
            (
                forecast.room_id,
                forecast.device_id,
                forecast.monitor_type,
                forecast.unit,
                _enum_value(forecast.horizon),
                forecast.predicted_at,
                forecast.predicted_value,
                _status_label(forecast.predicted_status),
                forecast.warning_threshold,
                forecast.control_threshold,
                forecast.method,
                forecast.sample_count,
                forecast.training_start,
                forecast.training_end,
                forecast.confidence,
                forecast.explanation,
                FORECAST_DISCLAIMER,
            )
        )
        _style_status_cell(
            worksheet.cell(row=worksheet.max_row, column=8),
            forecast.predicted_status,
        )
    if not forecasts:
        worksheet.append(
            (
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                FORECAST_DISCLAIMER,
            )
        )
    _finalize_table(
        worksheet,
        len(headers),
        datetime_columns=(6, 13, 14),
        numeric_columns=(7, 9, 10),
        integer_columns=(12,),
    )


def _write_cleaned_data_sheet(
    worksheet: Worksheet,
    records: Sequence[MonitoringRecord],
) -> None:
    headers = (
        "监测时间",
        "房间ID",
        "房间名称",
        "设备ID",
        "设备名称",
        "监测类型",
        "监测值",
        "单位",
        "预警值",
        "控制标准",
        "数据来源",
        "备注",
        "是否仅日期",
        "来源文件",
        "来源工作表",
        "来源行号",
        "导入顺序",
    )
    _write_headers(worksheet, headers)
    for record in records:
        worksheet.append(
            (
                record.monitored_at,
                record.room_id,
                record.room_name,
                record.device_id,
                record.device_name,
                record.monitor_type,
                record.value,
                record.unit,
                record.warning_threshold,
                record.control_threshold,
                record.data_source,
                record.note,
                "是" if record.date_only else "否",
                record.source_file,
                record.source_sheet,
                record.source_row,
                record.import_order,
            )
        )
    _finalize_table(
        worksheet,
        len(headers),
        datetime_columns=(1,),
        numeric_columns=(7, 9, 10),
        integer_columns=(16, 17),
    )


def _write_quality_issues_sheet(
    worksheet: Worksheet,
    issues: Sequence[QualityIssue],
) -> None:
    headers = (
        "级别",
        "问题代码",
        "问题信息",
        "来源文件",
        "来源工作表",
        "来源行号",
        "详情",
    )
    _write_headers(worksheet, headers)
    for issue in issues:
        worksheet.append(
            (
                issue.level,
                issue.code,
                issue.message,
                issue.source_file,
                issue.source_sheet,
                issue.source_row,
                _json_details(issue.details),
            )
        )
    _finalize_table(worksheet, len(headers), integer_columns=(6,))


def _write_headers(worksheet: Worksheet, headers: Sequence[str]) -> None:
    worksheet.append(tuple(headers))
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _finalize_table(
    worksheet: Worksheet,
    column_count: int,
    *,
    datetime_columns: Iterable[int] = (),
    numeric_columns: Iterable[int] = (),
    integer_columns: Iterable[int] = (),
    percent_columns: Iterable[int] = (),
) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(column_count)}{worksheet.max_row}"
    _set_widths(worksheet, _default_widths(column_count))

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in datetime_columns:
        for cell in worksheet.iter_cols(
            min_col=column,
            max_col=column,
            min_row=2,
            max_row=worksheet.max_row,
        ):
            for item in cell:
                item.number_format = DATETIME_FORMAT
    for column in numeric_columns:
        for cell in worksheet.iter_cols(
            min_col=column,
            max_col=column,
            min_row=2,
            max_row=worksheet.max_row,
        ):
            for item in cell:
                item.number_format = NUMBER_FORMAT
    for column in integer_columns:
        for cell in worksheet.iter_cols(
            min_col=column,
            max_col=column,
            min_row=2,
            max_row=worksheet.max_row,
        ):
            for item in cell:
                item.number_format = INTEGER_FORMAT
    for column in percent_columns:
        for cell in worksheet.iter_cols(
            min_col=column,
            max_col=column,
            min_row=2,
            max_row=worksheet.max_row,
        ):
            for item in cell:
                item.number_format = PERCENT_FORMAT


def _set_widths(worksheet: Worksheet, widths: Sequence[float]) -> None:
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width


def _default_widths(column_count: int) -> tuple[float, ...]:
    return tuple(16.0 for _ in range(column_count))


def _style_status_cell(cell, status: MonitoringStatus) -> None:
    cell.fill = STATUS_FILLS[status]
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _status_label(status: MonitoringStatus) -> str:
    return status.value


def _status_counts(statuses: Iterable[MonitoringStatus]) -> Counter[MonitoringStatus]:
    counts: Counter[MonitoringStatus] = Counter(statuses)
    for status in MonitoringStatus:
        counts.setdefault(status, 0)
    return counts


def _enum_value(value: Any) -> Any:
    return value.value if hasattr(value, "value") else value


def _join_reasons(reasons: Sequence[str]) -> str:
    return "；".join(reasons)


def _json_details(details: dict[str, Any]) -> str:
    return json.dumps(details, ensure_ascii=False, sort_keys=True)


def _source_columns(records: Sequence[Any]) -> tuple[str, str, str]:
    source_files = _unique_text(record.source_file for record in records)
    source_sheets = _unique_text(record.source_sheet for record in records)
    source_rows = _unique_text(str(record.source_row) for record in records)
    return source_files, source_sheets, source_rows


def _unique_text(values: Iterable[str]) -> str:
    return "、".join(dict.fromkeys(value for value in values if value))


def _top_risk_label(risks: Sequence[RiskResult]) -> str:
    if not risks:
        return ""
    risk = risks[0]
    return f"{risk.device_id}（{risk.score:.1f}）"


def _top_room_risk_label(risks: Sequence[RoomRiskResult]) -> str:
    if not risks:
        return ""
    risk = risks[0]
    return f"{risk.room_id}（{risk.score:.1f}）"
