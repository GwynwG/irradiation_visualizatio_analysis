from __future__ import annotations

from collections import defaultdict
from collections.abc import Mapping
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from io import BytesIO
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from .models import MonitoringRecord, QualityIssue
from .validation import is_finite_number, validate_room_device_ids


REQUIRED_COLUMNS = (
    "监测时间",
    "房间ID",
    "设备ID",
    "监测类型",
    "监测值",
    "单位",
    "预警值",
    "控制标准",
)
OPTIONAL_COLUMNS = ("房间名称", "设备名称", "数据来源", "备注")


@dataclass(frozen=True)
class UploadedWorkbook:
    filename: str
    content: bytes


@dataclass(frozen=True)
class ImportSummary:
    file_count: int
    raw_rows: int
    valid_rows: int
    blocked_rows: int
    exact_duplicate_rows: int
    conflict_keys: int
    room_count: int
    device_count: int
    monitor_types: tuple[str, ...]


@dataclass(frozen=True)
class ImportResult:
    records: list[MonitoringRecord]
    issues: list[QualityIssue]
    summary: ImportSummary
    candidate_sheets: dict[str, tuple[str, ...]]


@dataclass(frozen=True)
class _SheetCandidate:
    worksheet: Worksheet
    header_row: int
    columns: dict[str, int]


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_header(worksheet: Worksheet) -> _SheetCandidate | None:
    for row_number, cells in enumerate(
        worksheet.iter_rows(min_row=1, max_row=worksheet.max_row),
        start=1,
    ):
        columns: dict[str, int] = {}
        for column_index, cell in enumerate(cells):
            name = _text(cell.value)
            if name and name not in columns:
                columns[name] = column_index
        if all(name in columns for name in REQUIRED_COLUMNS):
            return _SheetCandidate(worksheet, row_number, columns)
    return None


def _number_format_has_time(number_format: str) -> bool:
    normalized = re.sub(r'"[^"]*"|\\.', "", number_format.lower())
    return "h" in normalized or "s" in normalized or "am/pm" in normalized


def _normalize_datetime(cell: Cell) -> tuple[datetime, bool] | None:
    value = cell.value

    if isinstance(value, datetime):
        date_only = not _number_format_has_time(cell.number_format or "")
        monitored_at = (
            datetime.combine(value.date(), time.min) if date_only else value
        )
        if monitored_at.tzinfo is not None:
            monitored_at = monitored_at.astimezone(timezone.utc).replace(tzinfo=None)
        return monitored_at, date_only

    if isinstance(value, date):
        return datetime.combine(value, time.min), True

    if not isinstance(value, str):
        return None

    text_value = value.strip()
    if not text_value:
        return None

    for date_format in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            parsed_date = datetime.strptime(text_value, date_format)
        except ValueError:
            continue
        return parsed_date, True

    try:
        parsed_datetime = datetime.fromisoformat(text_value.replace("Z", "+00:00"))
    except ValueError:
        return None

    if parsed_datetime.tzinfo is not None:
        parsed_datetime = parsed_datetime.astimezone(timezone.utc).replace(tzinfo=None)
    return parsed_datetime, False


def _parse_number(value: object) -> float | None:
    if isinstance(value, bool) or _is_blank(value):
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError, OverflowError):
        return None
    if not is_finite_number(parsed):
        return None
    return parsed


def _row_issue(
    *,
    code: str,
    message: str,
    filename: str,
    sheet_name: str,
    row_number: int,
    details: dict[str, Any] | None = None,
) -> QualityIssue:
    return QualityIssue(
        level="error",
        code=code,
        message=message,
        source_file=filename,
        source_sheet=sheet_name,
        source_row=row_number,
        details=details or {},
    )


def _record_signature(record: MonitoringRecord) -> tuple[object, ...]:
    return (
        record.date_only,
        record.value,
        record.unit,
        record.warning_threshold,
        record.control_threshold,
        record.room_name,
        record.device_name,
        record.data_source,
        record.note,
    )


def _source_details(record: MonitoringRecord) -> dict[str, object]:
    return {
        "file": record.source_file,
        "sheet": record.source_sheet,
        "row": record.source_row,
        "import_order": record.import_order,
    }


def _version_details(record: MonitoringRecord) -> dict[str, object]:
    return {
        "source_file": record.source_file,
        "source_sheet": record.source_sheet,
        "source_row": record.source_row,
        "import_order": record.import_order,
        "value": record.value,
        "unit": record.unit,
        "warning_threshold": record.warning_threshold,
        "control_threshold": record.control_threshold,
        "date_only": record.date_only,
        "room_name": record.room_name,
        "device_name": record.device_name,
        "data_source": record.data_source,
        "note": record.note,
    }


def _merge_records(
    records: list[MonitoringRecord],
) -> tuple[list[MonitoringRecord], list[QualityIssue], int, int]:
    grouped: dict[
        tuple[datetime, str, str, str],
        list[MonitoringRecord],
    ] = defaultdict(list)
    for record in records:
        grouped[record.key].append(record)

    merged: list[MonitoringRecord] = []
    issues: list[QualityIssue] = []
    exact_duplicate_rows = 0
    conflict_keys = 0

    for key_records in grouped.values():
        versions_by_signature: dict[tuple[object, ...], MonitoringRecord] = {}
        for record in key_records:
            signature = _record_signature(record)
            if signature in versions_by_signature:
                exact_duplicate_rows += 1
            versions_by_signature[signature] = record

        versions = sorted(
            versions_by_signature.values(),
            key=lambda record: record.import_order,
        )
        selected = versions[-1]
        merged.append(selected)

        if len(versions) > 1:
            conflict_keys += 1
            issues.append(
                QualityIssue(
                    level="warning",
                    code="conflicting_record",
                    message="Multiple versions share the same monitoring key; the latest version was selected.",
                    source_file=selected.source_file,
                    source_sheet=selected.source_sheet,
                    source_row=selected.source_row,
                    details={
                        "selected_source": _source_details(selected),
                        "versions": [
                            _version_details(version) for version in versions
                        ],
                    },
                )
            )

    return merged, issues, exact_duplicate_rows, conflict_keys


def _audit_history(records: list[MonitoringRecord]) -> list[QualityIssue]:
    grouped: dict[tuple[str, str], list[MonitoringRecord]] = defaultdict(list)
    for record in records:
        grouped[(record.device_id, record.monitor_type)].append(record)

    issues: list[QualityIssue] = []
    for (device_id, monitor_type), history in grouped.items():
        ordered = sorted(
            history,
            key=lambda record: (record.monitored_at, record.import_order),
        )
        latest = ordered[-1]
        units = tuple(dict.fromkeys(record.unit for record in ordered))
        thresholds = tuple(
            dict.fromkeys(
                (
                    record.warning_threshold,
                    record.control_threshold,
                )
                for record in ordered
            )
        )

        if len(units) > 1:
            issues.append(
                QualityIssue(
                    level="warning",
                    code="unit_changed",
                    message="The historical unit changed for this device and monitoring type.",
                    source_file=latest.source_file,
                    source_sheet=latest.source_sheet,
                    source_row=latest.source_row,
                    details={
                        "device_id": device_id,
                        "monitor_type": monitor_type,
                        "units": units,
                    },
                )
            )

        if len(thresholds) > 1:
            issues.append(
                QualityIssue(
                    level="warning",
                    code="threshold_changed",
                    message="The historical thresholds changed for this device and monitoring type.",
                    source_file=latest.source_file,
                    source_sheet=latest.source_sheet,
                    source_row=latest.source_row,
                    details={
                        "device_id": device_id,
                        "monitor_type": monitor_type,
                        "thresholds": thresholds,
                    },
                )
            )

    return issues


def import_workbooks(
    workbooks: list[UploadedWorkbook],
    selected_sheets: Mapping[str, str] | None = None,
) -> ImportResult:
    selected_sheets = selected_sheets or {}
    candidate_sheets: dict[str, tuple[str, ...]] = {}
    issues: list[QualityIssue] = []
    parsed_records: list[MonitoringRecord] = []
    raw_rows = 0
    valid_rows = 0
    blocked_rows = 0
    next_import_order = 0

    for uploaded_workbook in workbooks:
        try:
            workbook = load_workbook(
                BytesIO(uploaded_workbook.content),
                data_only=True,
                read_only=True,
            )
        except Exception as error:
            candidate_sheets[uploaded_workbook.filename] = ()
            issues.append(
                QualityIssue(
                    level="error",
                    code="workbook_read_error",
                    message="The workbook could not be read.",
                    source_file=uploaded_workbook.filename,
                    details={"error": str(error)},
                )
            )
            continue

        records_start = len(parsed_records)
        issues_start = len(issues)
        raw_rows_start = raw_rows
        valid_rows_start = valid_rows
        blocked_rows_start = blocked_rows
        import_order_start = next_import_order
        try:
            candidates = [
                candidate
                for worksheet in workbook.worksheets
                if (candidate := _find_header(worksheet)) is not None
            ]
            candidate_names = tuple(
                candidate.worksheet.title for candidate in candidates
            )
            candidate_sheets[uploaded_workbook.filename] = candidate_names

            if not candidates:
                issues.append(
                    QualityIssue(
                        level="error",
                        code="required_columns_missing",
                        message="No worksheet contains all required columns.",
                        source_file=uploaded_workbook.filename,
                        details={"required_columns": REQUIRED_COLUMNS},
                    )
                )
                continue

            selected_name = selected_sheets.get(uploaded_workbook.filename)
            if len(candidates) > 1:
                selected = next(
                    (
                        candidate
                        for candidate in candidates
                        if candidate.worksheet.title == selected_name
                    ),
                    None,
                )
                if selected is None:
                    issues.append(
                        QualityIssue(
                            level="error",
                            code="sheet_selection_required",
                            message="Select one candidate worksheet before importing.",
                            source_file=uploaded_workbook.filename,
                            details={"candidate_sheets": candidate_names},
                        )
                    )
                    continue
            else:
                selected = candidates[0]

            worksheet = selected.worksheet
            for row_number, cells in enumerate(
                worksheet.iter_rows(
                    min_row=selected.header_row + 1,
                    max_row=worksheet.max_row,
                ),
                start=selected.header_row + 1,
            ):
                if all(_is_blank(cell.value) for cell in cells):
                    continue

                raw_rows += 1
                values = {
                    name: cells[column_index].value
                    for name, column_index in selected.columns.items()
                    if column_index < len(cells)
                }
                missing_columns = [
                    name
                    for name in REQUIRED_COLUMNS
                    if _is_blank(values.get(name))
                ]
                if missing_columns:
                    blocked_rows += 1
                    issues.append(
                        _row_issue(
                            code="required_value_missing",
                            message="One or more required values are blank.",
                            filename=uploaded_workbook.filename,
                            sheet_name=worksheet.title,
                            row_number=row_number,
                            details={"columns": missing_columns},
                        )
                    )
                    continue

                monitored_at_cell = cells[selected.columns["监测时间"]]
                normalized_datetime = _normalize_datetime(monitored_at_cell)
                if normalized_datetime is None:
                    blocked_rows += 1
                    issues.append(
                        _row_issue(
                            code="invalid_monitored_at",
                            message="The monitoring time could not be parsed.",
                            filename=uploaded_workbook.filename,
                            sheet_name=worksheet.title,
                            row_number=row_number,
                            details={"value": values["监测时间"]},
                        )
                    )
                    continue
                monitored_at, date_only = normalized_datetime

                room_id = _text(values["房间ID"])
                device_id = _text(values["设备ID"])
                identifier_errors = validate_room_device_ids(room_id, device_id)
                if identifier_errors:
                    blocked_rows += 1
                    issues.append(
                        _row_issue(
                            code="invalid_identifier",
                            message="The room or device identifier is invalid.",
                            filename=uploaded_workbook.filename,
                            sheet_name=worksheet.title,
                            row_number=row_number,
                            details={"errors": identifier_errors},
                        )
                    )
                    continue

                number_specs = (
                    ("监测值", "invalid_value"),
                    ("预警值", "invalid_warning_threshold"),
                    ("控制标准", "invalid_control_threshold"),
                )
                parsed_numbers: dict[str, float] = {}
                invalid_number = False
                for column_name, code in number_specs:
                    parsed_number = _parse_number(values[column_name])
                    if parsed_number is None:
                        blocked_rows += 1
                        issues.append(
                            _row_issue(
                                code=code,
                                message=f"{column_name} must be a finite number.",
                                filename=uploaded_workbook.filename,
                                sheet_name=worksheet.title,
                                row_number=row_number,
                                details={"value": values[column_name]},
                            )
                        )
                        invalid_number = True
                        break
                    parsed_numbers[column_name] = parsed_number
                if invalid_number:
                    continue

                if parsed_numbers["预警值"] > parsed_numbers["控制标准"]:
                    blocked_rows += 1
                    issues.append(
                        _row_issue(
                            code="invalid_threshold_order",
                            message="The warning threshold cannot exceed the control threshold.",
                            filename=uploaded_workbook.filename,
                            sheet_name=worksheet.title,
                            row_number=row_number,
                            details={
                                "warning_threshold": parsed_numbers["预警值"],
                                "control_threshold": parsed_numbers["控制标准"],
                            },
                        )
                    )
                    continue

                parsed_records.append(
                    MonitoringRecord(
                        monitored_at=monitored_at,
                        date_only=date_only,
                        room_id=room_id,
                        device_id=device_id,
                        monitor_type=_text(values["监测类型"]),
                        value=parsed_numbers["监测值"],
                        unit=_text(values["单位"]),
                        warning_threshold=parsed_numbers["预警值"],
                        control_threshold=parsed_numbers["控制标准"],
                        source_file=uploaded_workbook.filename,
                        source_sheet=worksheet.title,
                        source_row=row_number,
                        import_order=next_import_order,
                        room_name=_text(values.get("房间名称")),
                        device_name=_text(values.get("设备名称")),
                        data_source=_text(values.get("数据来源")),
                        note=_text(values.get("备注")),
                    )
                )
                valid_rows += 1
                next_import_order += 1
        except Exception as error:
            del parsed_records[records_start:]
            del issues[issues_start:]
            raw_rows = raw_rows_start
            valid_rows = valid_rows_start
            blocked_rows = blocked_rows_start
            next_import_order = import_order_start
            candidate_sheets[uploaded_workbook.filename] = ()
            issues.append(
                QualityIssue(
                    level="error",
                    code="workbook_read_error",
                    message="The workbook could not be read.",
                    source_file=uploaded_workbook.filename,
                    details={"error": str(error)},
                )
            )
        finally:
            workbook.close()

    records, merge_issues, exact_duplicate_rows, conflict_keys = _merge_records(
        parsed_records
    )
    issues.extend(merge_issues)
    issues.extend(_audit_history(records))
    records.sort(
        key=lambda record: (
            record.monitored_at,
            record.room_id,
            record.device_id,
            record.monitor_type,
            record.import_order,
        )
    )

    summary = ImportSummary(
        file_count=len(workbooks),
        raw_rows=raw_rows,
        valid_rows=valid_rows,
        blocked_rows=blocked_rows,
        exact_duplicate_rows=exact_duplicate_rows,
        conflict_keys=conflict_keys,
        room_count=len({record.room_id for record in records}),
        device_count=len({record.device_id for record in records}),
        monitor_types=tuple(sorted({record.monitor_type for record in records})),
    )
    return ImportResult(records, issues, summary, candidate_sheets)
