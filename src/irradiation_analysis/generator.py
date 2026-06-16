from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
from io import BytesIO
from math import ceil, isfinite
import re
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from irradiation_analysis.excel_io import (
    OPTIONAL_COLUMNS,
    REQUIRED_COLUMNS,
    UploadedWorkbook,
)
from irradiation_analysis.snapshots import all_device_ids


DATA_SHEET_NAME = "监测数据"
THRESHOLD_SHEET_NAME = "设备阈值配置"
DATE_FORMAT = "yyyy-mm-dd hh:mm"
NUMBER_FORMAT = "0.000"
DEFAULT_ZIP_TIMESTAMP = (2026, 1, 1, 0, 0, 0)
CORE_PROPERTY_TIMESTAMP = b"2026-01-01T00:00:00Z"


@dataclass(frozen=True)
class MonitorTypeConfig:
    name: str
    unit: str
    warning_threshold: float
    control_threshold: float


_DEFAULT_MONITOR_TYPE_CONFIGS = (
    MonitorTypeConfig("γ剂量率", "μSv/h", 10.0, 20.0),
    MonitorTypeConfig("中子剂量率", "μSv/h", 5.0, 12.0),
)
DEFAULT_MONITOR_TYPES = tuple(config.name for config in _DEFAULT_MONITOR_TYPE_CONFIGS)


@dataclass(frozen=True)
class SimulationConfig:
    start: datetime
    end: datetime
    sampling_hours: int = 24
    monitor_types: tuple[MonitorTypeConfig, ...] = _DEFAULT_MONITOR_TYPE_CONFIGS
    output_mode: str = "single"
    warning_ratio: float = 0.05
    accident_ratio: float = 0.02
    rapid_growth_ratio: float = 0.02
    event_duration: int = 3
    seed: int | None = None


def build_blank_template() -> bytes:
    workbook, worksheet = _new_monitoring_workbook(_DEFAULT_MONITOR_TYPE_CONFIGS)
    worksheet.append([None] * len(_headers()))
    _apply_row_formats(worksheet, 2)
    _finalize_data_sheet(worksheet)
    return _save_workbook(workbook)


def build_prefilled_template(config: SimulationConfig | None = None) -> bytes:
    config = config or SimulationConfig(
        start=datetime(2026, 1, 1),
        end=datetime(2026, 1, 1),
    )
    _validate_config(config)

    workbook, worksheet = _new_monitoring_workbook(config.monitor_types)
    monitored_at = config.start
    for monitor_type in config.monitor_types:
        for device_id in all_device_ids():
            worksheet.append(
                _row(
                    monitored_at=monitored_at,
                    device_id=device_id,
                    monitor_type=monitor_type,
                    value=monitor_type.warning_threshold * 0.5,
                    data_source="template",
                    note="sample row",
                )
            )
            _apply_row_formats(worksheet, worksheet.max_row)

    _finalize_data_sheet(worksheet)
    return _save_workbook(workbook)


def generate_simulated_workbooks(config: SimulationConfig) -> list[UploadedWorkbook]:
    _validate_config(config)
    if config.output_mode != "single":
        raise ValueError(
            f"Unsupported output_mode {config.output_mode!r}; only 'single' is supported."
        )

    rng = np.random.default_rng(config.seed)
    times = _sampling_times(config)
    device_ids = all_device_ids()
    workbook, worksheet = _new_monitoring_workbook(config.monitor_types)
    scenarios = _scenario_assignments(
        series_count=len(device_ids) * len(config.monitor_types),
        config=config,
        rng=rng,
    )

    for time_index, monitored_at in enumerate(times):
        for monitor_index, monitor_type in enumerate(config.monitor_types):
            for device_index, device_id in enumerate(device_ids):
                series_index = monitor_index * len(device_ids) + device_index
                value = _simulated_value(
                    monitor_type=monitor_type,
                    scenario=scenarios.get(series_index, "normal"),
                    time_index=time_index,
                    sample_count=len(times),
                    event_duration=config.event_duration,
                    rng=rng,
                )
                worksheet.append(
                    _row(
                        monitored_at=monitored_at,
                        device_id=device_id,
                        monitor_type=monitor_type,
                        value=value,
                        data_source="simulated",
                    )
                )
                _apply_row_formats(worksheet, worksheet.max_row)

    _finalize_data_sheet(worksheet)
    filename = (
        f"simulated_monitoring_{config.start:%Y%m%d}_{config.end:%Y%m%d}.xlsx"
    )
    return [UploadedWorkbook(filename=filename, content=_save_workbook(workbook))]


def _headers() -> tuple[str, ...]:
    return REQUIRED_COLUMNS + OPTIONAL_COLUMNS


def _new_monitoring_workbook(
    monitor_types: tuple[MonitorTypeConfig, ...],
) -> tuple[Workbook, Worksheet]:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = DATA_SHEET_NAME
    _write_headers(worksheet, _headers())
    _add_threshold_sheet(workbook, monitor_types)
    return workbook, worksheet


def _write_headers(worksheet: Worksheet, headers: tuple[str, ...]) -> None:
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _add_threshold_sheet(
    workbook: Workbook,
    monitor_types: tuple[MonitorTypeConfig, ...],
) -> None:
    worksheet = workbook.create_sheet(THRESHOLD_SHEET_NAME)
    headers = ("监测类型", "单位", "预警值", "控制标准")
    _write_headers(worksheet, headers)
    for monitor_type in monitor_types:
        worksheet.append(
            [
                monitor_type.name,
                monitor_type.unit,
                monitor_type.warning_threshold,
                monitor_type.control_threshold,
            ]
        )
        worksheet.cell(row=worksheet.max_row, column=3).number_format = NUMBER_FORMAT
        worksheet.cell(row=worksheet.max_row, column=4).number_format = NUMBER_FORMAT
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column, width in enumerate((16, 12, 12, 12), start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width


def _finalize_data_sheet(worksheet: Worksheet) -> None:
    widths = (20, 10, 14, 16, 12, 12, 12, 12, 14, 16, 14, 24)
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def _apply_row_formats(worksheet: Worksheet, row_number: int) -> None:
    worksheet.cell(row=row_number, column=1).number_format = DATE_FORMAT
    for column in (5, 7, 8):
        worksheet.cell(row=row_number, column=column).number_format = NUMBER_FORMAT


def _row(
    *,
    monitored_at: datetime,
    device_id: str,
    monitor_type: MonitorTypeConfig,
    value: float,
    data_source: str,
    note: str = "",
) -> list[object]:
    room_id = device_id.split("-")[0]
    return [
        monitored_at,
        room_id,
        device_id,
        monitor_type.name,
        round(float(value), 6),
        monitor_type.unit,
        monitor_type.warning_threshold,
        monitor_type.control_threshold,
        f"Room {room_id}",
        f"Device {device_id}",
        data_source,
        note,
    ]


def _sampling_times(config: SimulationConfig) -> tuple[datetime, ...]:
    step = timedelta(hours=config.sampling_hours)
    times: list[datetime] = []
    current = config.start
    while current <= config.end:
        times.append(current)
        current += step
    return tuple(times)


def _scenario_assignments(
    *,
    series_count: int,
    config: SimulationConfig,
    rng: np.random.Generator,
) -> dict[int, str]:
    shuffled = list(rng.permutation(series_count))
    assignments: dict[int, str] = {}
    cursor = 0
    for scenario, ratio in (
        ("accident", config.accident_ratio),
        ("warning", config.warning_ratio),
        ("rapid", config.rapid_growth_ratio),
    ):
        count = min(_ratio_count(ratio, series_count), series_count - cursor)
        for series_index in shuffled[cursor : cursor + count]:
            assignments[int(series_index)] = scenario
        cursor += count
    return assignments


def _ratio_count(ratio: float, total: int) -> int:
    if ratio == 0:
        return 0
    return max(1, ceil(total * ratio))


def _simulated_value(
    *,
    monitor_type: MonitorTypeConfig,
    scenario: str,
    time_index: int,
    sample_count: int,
    event_duration: int,
    rng: np.random.Generator,
) -> float:
    warning = monitor_type.warning_threshold
    control = monitor_type.control_threshold
    event_start = max(0, sample_count - event_duration)

    if scenario == "accident" and time_index >= event_start:
        return control * rng.uniform(1.05, 1.35)

    if scenario == "warning" and time_index >= event_start:
        return warning + (control - warning) * rng.uniform(0.15, 0.75)

    if scenario == "rapid":
        return _rapid_growth_value(warning, time_index, sample_count)

    return max(0.001, warning * rng.uniform(0.20, 0.55))


def _rapid_growth_value(warning: float, time_index: int, sample_count: int) -> float:
    if sample_count <= 1:
        return warning * 0.35
    if time_index == sample_count - 1:
        return warning * 0.85
    return warning * (0.30 + 0.01 * time_index)


def _validate_config(config: SimulationConfig) -> None:
    if config.start > config.end:
        raise ValueError("start must be before or equal to end")
    if config.sampling_hours <= 0:
        raise ValueError("sampling_hours must be positive")
    if config.event_duration <= 0:
        raise ValueError("event_duration must be positive")
    for name, ratio in (
        ("warning_ratio", config.warning_ratio),
        ("accident_ratio", config.accident_ratio),
        ("rapid_growth_ratio", config.rapid_growth_ratio),
    ):
        if not isfinite(ratio) or ratio < 0 or ratio > 1:
            raise ValueError(f"{name} must be between 0 and 1")

    if not config.monitor_types:
        raise ValueError("monitor_types must not be empty")
    for monitor_type in config.monitor_types:
        if not monitor_type.name.strip():
            raise ValueError("monitor type name must not be blank")
        if not monitor_type.unit.strip():
            raise ValueError("monitor type unit must not be blank")
        if (
            not isfinite(monitor_type.warning_threshold)
            or not isfinite(monitor_type.control_threshold)
            or monitor_type.warning_threshold <= 0
            or monitor_type.control_threshold <= 0
            or monitor_type.warning_threshold >= monitor_type.control_threshold
        ):
            raise ValueError(
                "monitor type thresholds must be positive and warning must be below control"
            )


def _save_workbook(workbook: Workbook) -> bytes:
    workbook.properties.creator = "irradiation-analysis"
    workbook.properties.lastModifiedBy = "irradiation-analysis"
    workbook.properties.created = datetime(*DEFAULT_ZIP_TIMESTAMP[:6])
    workbook.properties.modified = datetime(*DEFAULT_ZIP_TIMESTAMP[:6])
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return _normalize_xlsx_zip(buffer.getvalue())


def _normalize_xlsx_zip(content: bytes) -> bytes:
    source = BytesIO(content)
    target = BytesIO()
    with ZipFile(source, "r") as input_zip:
        entries = [
            (info, input_zip.read(info.filename)) for info in input_zip.infolist()
        ]
    with ZipFile(target, "w", compression=ZIP_DEFLATED) as output_zip:
        for source_info, data in entries:
            target_info = ZipInfo(source_info.filename, DEFAULT_ZIP_TIMESTAMP)
            target_info.compress_type = source_info.compress_type or ZIP_DEFLATED
            target_info.external_attr = source_info.external_attr
            data = _normalize_core_properties(source_info.filename, data)
            output_zip.writestr(target_info, data)
    return target.getvalue()


def _normalize_core_properties(filename: str, data: bytes) -> bytes:
    if filename != "docProps/core.xml":
        return data
    return re.sub(
        rb"(<dcterms:(?:created|modified)[^>]*>)[^<]+(</dcterms:(?:created|modified)>)",
        lambda match: match.group(1) + CORE_PROPERTY_TIMESTAMP + match.group(2),
        data,
    )
